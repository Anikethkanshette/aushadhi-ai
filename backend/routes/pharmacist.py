from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from models import PharmacistLogin
from database import get_all_orders, load_medicines, update_order_status
from agents.notification_agent import NotificationAgent
from datetime import date
import os, io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

try:
    from google import genai
except ImportError:
    genai = None

router = APIRouter(prefix="/pharmacist", tags=["Pharmacist"])
notification_agent = NotificationAgent()

# ─── helpers ─────────────────────────────────────────────────────────────────

def _style_header(ws, headers: list[str], fill_hex: str = "1E3A5F"):
    """Write styled header row."""
    header_fill = PatternFill("solid", fgColor=fill_hex)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"),
        bottom=Side(style="thin", color="334155"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 22


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)


def _excel_response(wb: "openpyxl.Workbook", filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/login")
def pharmacist_login(credentials: PharmacistLogin):
    if credentials.username == "admin" and credentials.password == "admin":
        return {"status": "success", "data": {"access_token": "fake-pharmacist-jwt-token", "token_type": "bearer"}, "message": "Login successful", "error_code": None}
    raise HTTPException(status_code=401, detail="Invalid pharmacist credentials")

@router.get("/stats")
def get_dashboard_stats():
    orders = get_all_orders(force_refresh=True)
    medicines = load_medicines()
    
    pending_orders = [o for o in orders if o["status"].lower() == "pending"]
    low_stock = [m for m in medicines if int(m["stock_quantity"]) < int(m["min_stock_level"])]
    
    total_revenue = sum(float(o["total_amount"]) for o in orders if o["status"].lower() == "completed")
    
    return {
        "status": "success",
        "data": {
            "total_orders": len(orders),
            "pending_orders": len(pending_orders),
            "total_medicines": len(medicines),
            "low_stock_items": len(low_stock),
            "total_revenue": total_revenue
        },
        "message": "Dashboard stats retrieved successfully",
        "error_code": None
    }

@router.get("/orders")
def get_pharmacist_orders():
    orders = get_all_orders(force_refresh=True)
    # Return newest first
    return {
        "status": "success",
        "data": {
            "orders": sorted(orders, key=lambda x: x["purchase_date"], reverse=True)
        },
        "message": "Orders retrieved successfully",
        "error_code": None
    }

@router.put("/orders/{order_id}/status")
def update_status(order_id: str, status: str):
    valid_statuses = ["pending", "approved", "fulfilled", "rejected", "completed", "cancelled"]
    if status.lower() not in valid_statuses:
        raise HTTPException(status_code=400, detail="Invalid status")
        
    success = update_order_status(order_id, status.lower())
    if not success:
        raise HTTPException(status_code=404, detail="Order not found")
        
    # Trigger Notification Agent if order is fulfilled or approved
    if status.lower() in ["fulfilled", "completed", "approved"]:
        orders = get_all_orders()
        order = next((o for o in orders if o["order_id"] == order_id), None)
        if order:
            event_type = f"Order {status.capitalize()}"
            notif_copy = notification_agent.generate_and_dispatch(
                patient_id=order["patient_id"],
                abha_id=order["abha_id"],
                patient_name=order["patient_name"],
                event_type=event_type,
                details={"medicine_name": order["medicine_name"], "order_id": order_id}
            )
            return {
                "status": "success",
                "data": {"generated_notifications": notif_copy},
                "message": "Status updated successfully",
                "error_code": None
            }

    return {
        "status": "success",
        "data": {"message": "Status updated"},
        "message": "Status updated successfully",
        "error_code": None
    }


@router.put("/orders/bulk-status")
def bulk_update_status(body: dict):
    order_ids = body.get("order_ids", [])
    target_status = str(body.get("status", "")).lower()
    valid_statuses = ["pending", "approved", "fulfilled", "rejected", "completed", "cancelled"]

    if not isinstance(order_ids, list) or len(order_ids) == 0:
        raise HTTPException(status_code=400, detail="order_ids list is required")
    if target_status not in valid_statuses:
        raise HTTPException(status_code=400, detail="Invalid target status")

    updated = []
    not_found = []
    for order_id in order_ids:
        try:
            success = update_order_status(order_id, target_status)
            if success:
                updated.append(order_id)
            else:
                not_found.append(order_id)
        except Exception:
            not_found.append(order_id)

    return {
        "status": "success",
        "data": {
            "updated": updated,
            "not_found": not_found,
            "updated_count": len(updated),
        },
        "message": "Bulk status update completed",
        "error_code": None,
    }

@router.get("/inventory")
def get_inventory():
    medicines = load_medicines()
    return {
        "status": "success",
        "data": {
            "medicines": medicines
        },
        "message": "Inventory retrieved successfully",
        "error_code": None
    }


@router.post("/generate-po")
def generate_purchase_order():
    medicines = load_medicines()
    low_stock = [m for m in medicines if int(m["stock_quantity"]) < int(m["min_stock_level"])]
    
    if not low_stock:
        return {"po_draft": "All inventory levels are optimal. No restock needed."}
        
    po_items = []
    for m in low_stock:
        suggested_order = int(m["min_stock_level"]) * 2 - int(m["stock_quantity"])
        po_items.append(f"- {m['name']} (ID: {m['id']}), Current: {m['stock_quantity']}, Suggest Ordering: {suggested_order}")
        
    items_text = "\n".join(po_items)
    
    if genai:
        try:
            client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
            prompt = f"""
            You are drafting a formal Purchase Order (PO) to a pharmaceutical distributor (PharmaCorp India).
            Today's Date is {date.today()}.
            
            The pharmacy 'AushadhiAI Clinic' has the following low stock items that need immediate replenishment:
            {items_text}
            
            Please write a professional, well-formatted Purchase Order email/document.
            Include a table or list of the items, a request for delivery timelines, and standard PO formalities.
            """
            
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt
            )
            return {"po_draft": response.text}
        except Exception as e:
            print("Gemini PO Error:", e)

    fallback_po = (
        f"PURCHASE ORDER\n"
        f"Date: {date.today()}\n"
        f"To: PharmaCorp India\n\n"
        f"Please supply the following items:\n{items_text}\n\n"
        f"Requested Delivery: ASAP"
    )
    return {"po_draft": fallback_po}


# ─── Patient Notifications ────────────────────────────────────────────────────

# In-memory store for demo (persists for session)
_NOTIFICATIONS: list[dict] = []

@router.post("/notify-patient")
def notify_patient(body: dict):
    """Pharmacist sends a custom notification to a patient."""
    patient_id   = body.get("patient_id", "")
    patient_name = body.get("patient_name", "")
    subject      = body.get("subject", "")
    message      = body.get("message", "")
    notif_type   = body.get("type", "general")  # general | refill | alert | appointment

    if not patient_name or not message:
        raise HTTPException(status_code=400, detail="patient_name and message are required")

    # Try to enhance message with Gemini
    enhanced = message
    if genai and os.getenv("GEMINI_API_KEY"):
        try:
            client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
            prompt = (
                f"You are a friendly pharmacist assistant at AushadhiAI pharmacy.\n"
                f"Rewrite the following notification message to make it warm, professional, and clear.\n"
                f"Patient name: {patient_name}\n"
                f"Notification type: {notif_type}\n"
                f"Original message: {message}\n"
                f"Reply with ONLY the improved message, no extra commentary."
            )
            resp = client.models.generate_content(model='gemini-2.0-flash', contents=prompt)
            enhanced = resp.text.strip()
        except Exception as e:
            print("Gemini enhance error:", e)

    notif = {
        "id":           f"NOTIF-{len(_NOTIFICATIONS)+1:04d}",
        "patient_id":   patient_id,
        "patient_name": patient_name,
        "subject":      subject or f"Message from your Pharmacist",
        "message":      message,
        "enhanced":     enhanced,
        "type":         notif_type,
        "sent_at":      str(date.today()),
        "read":         False,
    }
    _NOTIFICATIONS.append(notif)
    return {"success": True, "notification": notif}


@router.get("/notifications")
def get_notifications(notif_type: str = None, read: bool = None):
    """Get all sent notifications (pharmacist view)."""
    notifications = list(reversed(_NOTIFICATIONS))
    if notif_type:
        notifications = [n for n in notifications if str(n.get("type", "")).lower() == notif_type.lower()]
    if read is not None:
        notifications = [n for n in notifications if bool(n.get("read", False)) == read]
    return {
        "status": "success",
        "data": {
            "notifications": notifications
        },
        "message": "Notifications retrieved successfully",
        "error_code": None
    }


@router.post("/alerts/low-stock-scan")
def generate_low_stock_alerts():
    """Auto-generate pharmacist low-stock alert notifications for today's low inventory items."""
    medicines = load_medicines()
    today_str = str(date.today())

    low_stock_items = [
        m for m in medicines
        if int(m.get("stock_quantity", 0)) < int(m.get("min_stock_level", 0))
    ]

    created = []
    for item in low_stock_items:
        medicine_id = item.get("id")
        already_exists = any(
            n.get("type") == "low_stock"
            and n.get("sent_at") == today_str
            and n.get("medicine_id") == medicine_id
            for n in _NOTIFICATIONS
        )
        if already_exists:
            continue

        notif = {
            "id": f"NOTIF-{len(_NOTIFICATIONS)+1:04d}",
            "patient_id": "",
            "patient_name": "Pharmacist Team",
            "subject": f"Low stock alert: {item.get('name', medicine_id)}",
            "message": f"{item.get('name', medicine_id)} is below minimum stock ({item.get('stock_quantity')} < {item.get('min_stock_level')}).",
            "enhanced": f"Action needed: reorder {item.get('name', medicine_id)}. Current stock {item.get('stock_quantity')} {item.get('unit', 'units')}.",
            "type": "low_stock",
            "medicine_id": medicine_id,
            "sent_at": today_str,
            "read": False,
        }
        _NOTIFICATIONS.append(notif)
        created.append(notif)

    return {
        "status": "success",
        "data": {
            "created": created,
            "created_count": len(created),
            "low_stock_count": len(low_stock_items),
        },
        "message": "Low-stock alerts processed successfully",
        "error_code": None,
    }


@router.get("/patient-notifications/{patient_id}")
def get_patient_notifications(patient_id: str):
    """Get notifications for a specific patient."""
    patient_notifs = [n for n in _NOTIFICATIONS if n["patient_id"] == patient_id or not n["patient_id"]]
    return {
        "status": "success",
        "data": {
            "notifications": list(reversed(patient_notifs))
        },
        "message": "Patient notifications retrieved successfully",
        "error_code": None
    }


# ─── Excel Exports ──────────────────────────────────────────────────────────

@router.get("/export/orders")
def export_orders_excel():
    if not XLSX_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    orders = get_all_orders(force_refresh=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order History"

    headers = ["Order ID", "Patient ID", "Patient Name", "ABHA ID", "Medicine ID",
               "Medicine Name", "Qty", "Dosage Frequency", "Purchase Date",
               "Next Refill Date", "Total Amount (₹)", "Status", "Tx ID"]
    _style_header(ws, headers, "1E3A5F")

    alt_fill = PatternFill("solid", fgColor="0F172A")
    for i, o in enumerate(orders, 2):
        row = [
            o.get("order_id", ""), o.get("patient_id", ""), o.get("patient_name", ""),
            o.get("abha_id", ""), o.get("medicine_id", ""), o.get("medicine_name", ""),
            o.get("quantity", ""), o.get("dosage_frequency", ""),
            o.get("purchase_date", ""), o.get("next_refill_date", ""),
            float(o.get("total_amount", 0)), o.get("status", ""), o.get("tx_id", ""),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(color="E2E8F0", size=10)
            if i % 2 == 0:
                cell.fill = alt_fill
    _auto_width(ws)
    ws.freeze_panes = "A2"
    return _excel_response(wb, f"AushadhiAI_Orders_{date.today()}.xlsx")


@router.get("/export/products")
def export_products_excel():
    if not XLSX_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    medicines = load_medicines()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Catalogue"

    headers = ["Product ID", "Name", "Generic Name", "Category", "Stock Qty",
               "Unit", "Price (₹)", "Rx Required", "Min Stock Level", "Supplier"]
    _style_header(ws, headers, "064E3B")

    low_fill = PatternFill("solid", fgColor="450A0A")
    alt_fill = PatternFill("solid", fgColor="0F172A")
    for i, m in enumerate(medicines, 2):
        is_low = int(m.get("stock_quantity", 0)) < int(m.get("min_stock_level", 0))
        row = [
            m.get("id", ""), m.get("name", ""), m.get("generic_name", ""),
            m.get("category", ""), int(m.get("stock_quantity", 0)), m.get("unit", ""),
            float(m.get("price", 0)), str(m.get("prescription_required", "false")).title(),
            int(m.get("min_stock_level", 0)), m.get("supplier", ""),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(color="E2E8F0", size=10)
            cell.fill = low_fill if is_low else (alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="0D1117"))
    _auto_width(ws)
    ws.freeze_panes = "A2"
    return _excel_response(wb, f"AushadhiAI_Products_{date.today()}.xlsx")


@router.get("/export/patients")
def export_patients_excel():
    if not XLSX_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    # Import inline to avoid circular
    from routes.patients import SIMULATED_PATIENTS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patient Registry"

    headers = ["Patient ID", "ABHA ID", "Name", "Age", "Gender", "Blood Group",
               "Height (cm)", "Weight (kg)", "Phone", "Email",
               "Chronic Conditions", "Current Medications", "Allergies", "Last Visit"]
    _style_header(ws, headers, "3B0764")

    alt_fill = PatternFill("solid", fgColor="0F172A")
    for i, p in enumerate(SIMULATED_PATIENTS.values(), 2):
        row = [
            p.get("patient_id", ""), p.get("abha_id", ""), p.get("name", ""),
            p.get("age", ""), p.get("gender", ""), p.get("blood_group", ""),
            p.get("height_cm", ""), p.get("weight_kg", ""),
            p.get("phone", ""), p.get("email", ""),
            ", ".join(p.get("chronic_conditions", [])),
            ", ".join(p.get("current_medications", [])),
            ", ".join(p.get("allergies", [])),
            p.get("last_visit", ""),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(color="E2E8F0", size=10)
            if i % 2 == 0:
                cell.fill = alt_fill
    _auto_width(ws)
    ws.freeze_panes = "A2"
    return _excel_response(wb, f"AushadhiAI_Patients_{date.today()}.xlsx")
